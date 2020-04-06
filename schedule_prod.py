import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Border, Side
from openpyxl.styles import colors


class Schedule:
    def __init__(self, class_name, class_num, sched_weeks, date):
        self.class_name = class_name
        self.class_num = class_num
        self.sched_weeks = sched_weeks
        self.date = date

        # open a workbook
        self.wb = Workbook()
        self.ws = self.wb.active

        # Add Class Name, number of classes, and month to the top of sheet
        self.make_header(self.date, self.class_name, self.class_num)

        # merge cells that say the day and then print the day and date
        self.merge_cells(date, sched_weeks)
        self.set_width()
        self.print_dates(self.sched_weeks, self.date)

    def make_header(self, date, class_name, num_of_classes):
        month = date.strftime('%B')
        input3 = f"{num_of_classes} Classes"

        self.ws['J1'] = class_name.title()
        self.ws['J2'] = month
        self.ws['J3'] = input3

    def merge_cells(self, date, sched_weeks):
        self.ws.merge_cells('C2:H3')
        cell = self.ws['c2']
        cell.value = f"Turbo English - {date.strftime('%B')}"
        ft = Font(name='Calibri', size=24)
        cell.font = ft
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for x in range(sched_weeks):
            for y in range(10):
                if y % 2 != 0:
                    row = (x + 5) + (x * 2)
                    self.ws.merge_cells(start_row=row, start_column=y, end_row=row, end_column=y + 1)
                    self.ws.merge_cells(start_row=row + 1, start_column=y, end_row=row + 1, end_column=y + 1)

    def set_width(self):
        self.ws.column_dimensions['A'].width = 17
        self.ws.column_dimensions['B'].width = 17
        self.ws.column_dimensions['C'].width = 17
        self.ws.column_dimensions['D'].width = 17
        self.ws.column_dimensions['E'].width = 17
        self.ws.column_dimensions['F'].width = 17
        self.ws.column_dimensions['G'].width = 17
        self.ws.column_dimensions['H'].width = 17
        self.ws.column_dimensions['I'].width = 17
        self.ws.column_dimensions['J'].width = 17
        self.ws.row_dimensions[7].height = 88.5
        self.ws.row_dimensions[10].height = 88.5
        self.ws.row_dimensions[13].height = 88.5
        self.ws.row_dimensions[16].height = 88.5
        self.ws.row_dimensions[19].height = 88.5



    def print_dates(self, sched_weeks, date):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style="thin"),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for x in range(sched_weeks):
            for y in range(10):
                if y % 2 == 0:
                    day = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
                    row = (x + 5) + (x * 2)
                    col = y + 1

                    # print the day of the week
                    c = self.ws.cell(row, col)
                    index = int(y / 2)
                    c.value = day[index]
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border = thin_border
                    c.offset(0, 1).border = thin_border

                    # print the date
                    d = self.ws.cell(row + 1, col)
                    new_date = date
                    delta = (x * 7) + (y / 2)
                    new_date += datetime.timedelta(days=delta)
                    d.value = new_date.strftime("%m") + "/" + new_date.strftime("%d")
                    d.alignment = Alignment(horizontal="center", vertical="center")
                    d.fill = PatternFill(fill_type='solid', start_color='C5CACD', end_color='828282')
                    d.border = thin_border
                    d.offset(0, 1).border = thin_border

    def print_to_cell(self, string, row, col, color="none"):
        d = self.ws.cell(row, col)
        d.alignment = Alignment(vertical='top', wrap_text=True)
        d.value = string
        d.border = Border(left=Side(style='thin'),
                          right=Side(style="thin"),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
        if color == 'cls_todo':
            d.fill = PatternFill(fill_type='solid', start_color='D1D5D7', end_color='D1D5D7')
        elif color == 'hw_todo':
            d.fill = PatternFill(fill_type='solid', start_color='E1E4E5', end_color='E1E4E5')
        elif color == 'no_school':
            d.fill = PatternFill(fill_type='solid', start_color='E06666', end_color='E06666')

    def loop_through_weeks(self, sched_weeks, book_set, todo_func):
        for x in range(sched_weeks):
            for y in range(10):
                if y % 2 == 0:
                    # The original msg string has both class and homework, separated by a comma
                    # I split them here for two strings, and will write them at the same time.
                    classTodo = todo_func.split(",")[0]
                    hwTodo = todo_func.split(",")[1]

                    # We are now in the correct cell
                    row = (x+7)+(x*2)
                    self.print_to_cell(classTodo, row, y+1)
                    self.print_to_cell(hwTodo, row, y+2)

                else:
                    break
        return None

    def save_book(self, string):
        fileName = string + ".xlsx"
        self.wb.save(fileName)
